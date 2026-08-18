# -*- coding: utf-8 -*-
# =============================================================================
#  行业 / 指数 温度计 —— 单文件版（可直接粘进 Jupyter 一个 cell 运行）
# -----------------------------------------------------------------------------
#  逆向自 Excel 模板《行业温度计_量价_估值》：
#     温度 = Σ(各指标滚动分位数 × 权重) / Σ权重
#  - 实盘取数：iFinD SDK(iFinDPy) 的 THS_DS（见 run_batch / fetch_index）
#  - 离线演算：用模板自带缓存数据复算（见 demo_offline），无需 iFinD
#
#  依赖: pip install pandas numpy openpyxl  (实盘再加 iFinDPy)
# =============================================================================
import os
import numpy as np
import pandas as pd

# #############################################################################
#  ★★★ 用户配置区：在办公电脑上只需改这一块 ★★★
# -----------------------------------------------------------------------------
#  Windows 路径请用「原始字符串」即前面加 r，例如  r"D:\工作\行业温度计"
#  （这样反斜杠 \ 不会被转义）。Mac/Linux 用 "/Users/xxx/行业温度计" 即可。
# #############################################################################

# 工作根目录：缓存 / 输出 / 模板 默认都放这里。
#   留空 "" = 用「当前工作目录」(即你启动 Jupyter / 运行脚本的那个文件夹)。
BASE_DIR = r"D:\行业温度计"          # ← 改成你办公电脑放文件的文件夹；不确定就先填 ""

# 下面三个一般留空即可（=自动放在 BASE_DIR 下）。想单独指定就填完整绝对路径。
CACHE_DIR   = ""    # 缓存目录    ，空 = BASE_DIR\ifind_cache
OUTPUT_PATH = ""    # 输出Excel   ，空 = BASE_DIR\指数温度_汇总.xlsx
XLSX_PATH   = ""    # 模板(仅离线演算 demo_offline 用)，空 = BASE_DIR\转换 行业温度计_量价_估值.xlsx

# iFinD 账号：留空则读环境变量 IFIND_USER / IFIND_PWD；也可直接写死在这里。
IFIND_USER = ""
IFIND_PWD  = ""

# 取数 / 滚动参数（一般不用动）
DATA_BEGIN   = "2007-01-01"   # 取数起点（PE/PB 需 750 日滚动 + 250 日分位，要够长）
MARKET_CODE  = "700001.TI"    # 全市场(万得全A)，用于"拥挤度""超额PB"
OVERLAP_DAYS = 7              # 增量刷新时回补最近 N 天(防数据修订/补登)，再拼到缓存上

# 截止日期（取到哪天为止）。日中跑时当天数据还没出，应取到「上一日」。
#   ""        = 自动取 今天-1 天（与模板 E2=TODAY()-1 一致，推荐日中跑用这个）
#   "today"   = 取到今天（收盘后跑、确认当天数据已出时用）
#   "2026-05-28" = 写死某个具体日期（补算历史某天时用）
END_DATE = ""

# ---- 以下自动把上面的相对/留空路径解析成绝对路径，无需修改 ----
_BASE = BASE_DIR if BASE_DIR else os.getcwd()
CACHE_DIR   = CACHE_DIR   if CACHE_DIR   else os.path.join(_BASE, "ifind_cache")
OUTPUT_PATH = OUTPUT_PATH if OUTPUT_PATH else os.path.join(_BASE, "指数温度_汇总.xlsx")
XLSX_PATH   = XLSX_PATH   if XLSX_PATH   else os.path.join(_BASE, "转换 行业温度计_量价_估值.xlsx")
os.makedirs(_BASE, exist_ok=True)

# =============================================================================
#  ① 指标 / 权重配置  —— 要改权重 / 阈值 / 指标id，动这一块
# =============================================================================

# iFinD 数据接口（id 与 Excel thsiFinD 完全一致；indiparams=THS_DS 指标参数）
IFIND_INDICATORS = {
    "close":     ("ths_close_price_stock",              "102,"),
    "amt":       ("ths_amt_stock",                      ""),
    "turnover":  ("ths_turnover_ratio_stock",           ""),
    "raise_num": ("ths_constituent_raise_number_index", ""),
    "total_num": ("ths_constituent_num_index",          ""),
    "limit_up":  ("ths_constituent_up_number_index",    ""),   # 原模板id含空格,已去掉
    "limit_dn":  ("ths_constituent_dl_number_index",    ""),
    "pe":        ("ths_pe_ttm_sr_index",                "100,100"),
    "pb":        ("ths_pb_index",                       "108,100"),
}
IFIND_MARKET_INDICATORS = {
    "mkt_amt": ("ths_amt_stock", ""),
    "mkt_pb":  ("ths_pb_index",  "108,100"),
}

# 权重（模板 H3:R3）：超额PB=0，其余=1；创新高(S)不进公式
WEIGHTS = {
    "月跌幅分位": 1, "成交额增速分位": 1, "换手率分位": 1, "上涨占比分位": 1,
    "偏离年线分位": 1, "涨停跌停分位": 1, "平均成交额分位": 1,
    "PE分位": 1, "PB分位": 1, "超额PB分位": 0, "拥挤度分位": 1,
}

# 各底稿滚动窗口 / 行偏移（与 Excel 公式精确对应）
WIN = {
    "ret_offset": 29, "ret_div": 30, "ret_pct_win": 250,
    "amt_avg_win": 30, "amtgrow_offset": 30, "amtgrow_pct_win": 250,
    "turn_avg_win": 30, "turn_pct_win": 250,
    "raise_avg_win": 30, "raise_pct_win": 250,
    "ma_win": 250, "dev_pct_win": 250,
    "limit_avg_win": 30, "limit_pct_win": 250,
    "avgamt_avg_win": 30, "avgamt_pct_win": 250,
    "crowd_pct_win": 250,
    "pe_pct_win": 750, "pb_pct_win": 750, "expb_pct_win": 750,
}

# -----------------------------------------------------------------------------
#  乖离率（刘晨明/广发策略，减法版）配置  —— 独立计算，不并入温度
#    乖离率 = (ln(close) - EMA(ln(close),20)) * 100
#    分档：>15 过热 / 5~15 强势 / -5~5 偏弱(按需求统一归档) / <-5 止损
#    趋势方向：近20日收盘价在均线上方(=乖离率>0)的天数 ≥10 → 上行，否则 转弱
# -----------------------------------------------------------------------------
BIAS_EMA_SPAN   = 20     # EMA 周期（对 ln(close)）
BIAS_TREND_WIN  = 20     # 趋势判定回看天数
BIAS_TREND_MIN  = 10     # 近 BIAS_TREND_WIN 日中 乖离率>0 至少 N 天才算"上行"
BIAS_BANDS = [           # (下限含, 上限不含, 档位名)；从高到低
    (15,    np.inf, "过热"),
    (5,     15,     "强势"),
    (-5,    5,      "偏弱"),   # ← 原文 -5~0 / 0~5 等细分，按需求统一为"偏弱"
    (-np.inf, -5,   "止损"),
]


# =============================================================================
#  ② 计算引擎  —— 数组一律【按日期降序】(index0=最新), 与底稿行顺序一致
# =============================================================================
def percentrank_inc(arr, x):
    """精确复刻 Excel _xlfn.PERCENTRANK.INC(arr, x) → 0~1。"""
    a = np.sort(np.asarray([v for v in arr if v is not None and not np.isnan(v)],
                           dtype=float))
    n = a.size
    if n == 0 or x is None or (isinstance(x, float) and np.isnan(x)):
        return np.nan
    if n == 1:
        return 0.0
    if x <= a[0]:
        return 0.0
    if x >= a[-1]:
        return 1.0
    idx = np.searchsorted(a, x, side="left")
    if a[idx] == x:
        return idx / (n - 1)
    lo = idx - 1
    return (lo + (x - a[lo]) / (a[idx] - a[lo])) / (n - 1)


def roll_pctrank(series, win):
    """out[i] = PERCENTRANK.INC(series[i:i+win], series[i])（回看过去 win 天）。"""
    s = np.asarray(series, dtype=float)
    out = np.full(s.size, np.nan)
    for i in range(s.size):
        w = s[i:i + win]
        if np.count_nonzero(~np.isnan(w)) >= 2:
            out[i] = percentrank_inc(w, s[i])
    return out


def roll_mean(series, win):
    """降序数组上的尾窗均值 = mean(series[i:i+win])，忽略 NaN。"""
    s = pd.Series(series, dtype=float)
    return s[::-1].rolling(win, min_periods=1).mean()[::-1].to_numpy()


def shift_back(series, offset):
    """series[i+offset]（更早 offset 天），越界给 NaN。"""
    s = np.asarray(series, dtype=float)
    out = np.full(s.size, np.nan)
    if offset < s.size:
        out[:s.size - offset] = s[offset:]
    return out


# ----- 各指标 → 最终分位数列（入参 d: 原始降序序列 dict） -----
def calc_month_drop(d, W):
    c = d["close"]; base = shift_back(c, W["ret_offset"])
    with np.errstate(divide="ignore", invalid="ignore"):
        ret = ((c - base) / base) / W["ret_div"]
    return roll_pctrank(ret, W["ret_pct_win"])

def calc_amt_growth(d, W):
    avg = roll_mean(d["amt"] / 1e8, W["amt_avg_win"])
    base = shift_back(avg, W["amtgrow_offset"])
    with np.errstate(divide="ignore", invalid="ignore"):
        grow = (avg - base) / base
    return roll_pctrank(grow, W["amtgrow_pct_win"])

def calc_turnover(d, W):
    return roll_pctrank(roll_mean(d["turnover"], W["turn_avg_win"]), W["turn_pct_win"])

def calc_raise_ratio(d, W):
    with np.errstate(divide="ignore", invalid="ignore"):
        ratio = d["raise_num"] / d["total_num"]
    return roll_pctrank(roll_mean(ratio, W["raise_avg_win"]), W["raise_pct_win"])

def calc_year_line_dev(d, W):
    dev = d["close"] - roll_mean(d["close"], W["ma_win"])
    return roll_pctrank(dev, W["dev_pct_win"])

def calc_limit(d, W):
    avg = roll_mean(d["limit_up"] - d["limit_dn"], W["limit_avg_win"])
    return roll_pctrank(avg, W["limit_pct_win"])

def calc_avg_amt(d, W):
    avg = roll_mean(d["amt"] / 1e8, W["avgamt_avg_win"])
    return roll_pctrank(avg, W["avgamt_pct_win"])

def calc_pe(d, W):
    return roll_pctrank(d["pe"], W["pe_pct_win"])

def calc_pb(d, W):
    return roll_pctrank(d["pb"], W["pb_pct_win"])

def calc_excess_pb(d, W):
    if "mkt_pb" not in d: return np.full(d["pb"].size, np.nan)
    return roll_pctrank(d["pb"] - d["mkt_pb"], W["expb_pct_win"])

def calc_crowding(d, W):
    if "mkt_amt" not in d: return np.full(d["amt"].size, np.nan)
    with np.errstate(divide="ignore", invalid="ignore"):
        ratio = (d["amt"] / 1e8) / (d["mkt_amt"] / 1e8)
    return roll_pctrank(ratio, W["crowd_pct_win"])

INDICATOR_FUNCS = {
    "月跌幅分位": calc_month_drop, "成交额增速分位": calc_amt_growth,
    "换手率分位": calc_turnover, "上涨占比分位": calc_raise_ratio,
    "偏离年线分位": calc_year_line_dev, "涨停跌停分位": calc_limit,
    "平均成交额分位": calc_avg_amt, "PE分位": calc_pe, "PB分位": calc_pb,
    "超额PB分位": calc_excess_pb, "拥挤度分位": calc_crowding,
}


# ----- 乖离率（减法版）独立计算：仅依赖 close，不并入温度 -----
def calc_bias(close_desc, span=BIAS_EMA_SPAN):
    """减法版乖离率(%)：(ln(close)-EMA(ln(close),span))*100。入/出均为降序数组。"""
    asc = np.asarray(close_desc, dtype=float)[::-1]           # 转升序算 EMA
    lnc = pd.Series(np.log(asc))
    ema = lnc.ewm(span=span, adjust=False).mean()             # 通达信 EMA
    bias = (lnc - ema) * 100.0
    return bias.to_numpy()[::-1]                              # 转回降序

def classify_bias(value):
    """把乖离率数值映射到档位名。"""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "缺失"
    for lo, hi, name in BIAS_BANDS:
        if lo <= value < hi:
            return name
    return "缺失"

def trend_direction(bias_desc, win=BIAS_TREND_WIN, min_above=BIAS_TREND_MIN):
    """趋势方向：近 win 日 乖离率>0(=收盘在均线上方) 的天数 ≥min_above → '上行'，否则 '转弱'。
       返回与日期等长的标签数组(降序)。"""
    b = np.asarray(bias_desc, dtype=float)
    out = np.empty(b.size, dtype=object)
    for i in range(b.size):
        w = b[i:i + win]
        w = w[~np.isnan(w)]
        if w.size == 0:
            out[i] = "缺失"
        else:
            out[i] = "上行" if np.count_nonzero(w > 0) >= min_above else "转弱"
    return out


def compute_temperature(data, weights=WEIGHTS, win=WIN):
    """data: dict(含降序 'dates' 及各原始序列) → DataFrame(日期+各分位+温度, 降序)。"""
    out = pd.DataFrame({"日期": data["dates"]})
    for name in weights:
        out[name] = INDICATOR_FUNCS[name](data, win)
    active = [k for k, w in weights.items() if w != 0]
    wsum = sum(weights[k] for k in active)
    acc = np.zeros(len(out)); valid = np.ones(len(out), dtype=bool)
    for k in active:
        col = out[k].to_numpy(dtype=float)
        acc += np.nan_to_num(col) * weights[k]
        valid &= ~np.isnan(col)
    out["温度"] = np.where(valid, acc / wsum, np.nan)

    # —— 乖离率 & 趋势方向：独立计算，不并入温度 ——
    if "close" in data:
        bias = calc_bias(data["close"])
        out["乖离率%"] = np.round(bias, 2)
        out["乖离率分档"] = [classify_bias(v) for v in bias]
        out["趋势方向"] = trend_direction(bias)
    return out


# =============================================================================
#  ③ 数据层（实盘 iFinD + 本地增量缓存）—— iFinDPy 仅在真正取数时才 import
#     缓存策略：每个代码一份 CSV(原始序列,按日升序)。
#       · 首次见到的代码 → 拉完整区间 DATA_BEGIN..今天
#       · 已有缓存 → 只拉「上次最后一天 - OVERLAP_DAYS」..今天，再拼接覆盖
#     这样每次刷新只消耗"新增交易日 × 指标数"的接口量。
# =============================================================================
_LOGGED_IN = False

def ifind_login(user=None, pwd=None):
    global _LOGGED_IN
    if _LOGGED_IN: return
    from iFinDPy import THS_iFinDLogin
    ret = THS_iFinDLogin(user or IFIND_USER or os.getenv("IFIND_USER", ""),
                         pwd or IFIND_PWD or os.getenv("IFIND_PWD", ""))
    if ret not in (0, "0"):
        raise RuntimeError(f"iFinD 登录失败, 返回码={ret}")
    _LOGGED_IN = True

def _ds(code, indicator, indiparams, begin, end):
    from iFinDPy import THS_DS
    r = THS_DS(code, indicator, indiparams, "", begin, end)
    df = getattr(r, "data", r)
    if df is None or len(df) == 0:
        return pd.Series(dtype=float)
    df = pd.DataFrame(df)
    val_col = [c for c in df.columns if c not in ("thscode", "time")][-1]
    s = pd.Series(df[val_col].values, index=pd.to_datetime(df["time"].values))
    return pd.to_numeric(s, errors="coerce")

# ---------- 本地缓存读写 ----------
def _cache_path(code):
    os.makedirs(CACHE_DIR, exist_ok=True)
    return os.path.join(CACHE_DIR, code.replace("/", "_") + ".csv")

def _load_raw_cache(code):
    p = _cache_path(code)
    if not os.path.exists(p):
        return None
    df = pd.read_csv(p, index_col=0, parse_dates=True, encoding="utf-8-sig")
    return df.sort_index()

def _save_raw_cache(code, df):
    df.sort_index().to_csv(_cache_path(code), encoding="utf-8-sig")

def _fetch_raw_block(code, indicators, begin, end):
    """从 iFinD 拉一段区间的所有原始指标，返回 DataFrame(按日升序, 列=指标键)。"""
    cols = {k: _ds(code, ind, par, begin, end)
            for k, (ind, par) in indicators.items()}
    df = pd.DataFrame(cols)
    if not df.empty:
        df.index = pd.to_datetime(df.index)
        df = df.sort_index()
    return df

def _combine_cache(old, new):
    if old is None or old.empty: return new
    if new is None or new.empty: return old
    keep = old[~old.index.isin(new.index)]          # 重叠日用新数据覆盖
    return pd.concat([keep, new]).sort_index()

def _default_end():
    """根据配置 END_DATE 算默认截止日（日中跑默认取上一日）。"""
    if not END_DATE:                       # 空 → 今天-1（与模板 E2=TODAY()-1 一致）
        return (pd.Timestamp.today() - pd.Timedelta(days=1)).strftime("%Y-%m-%d")
    if END_DATE.lower() == "today":        # 收盘后跑可取到今天
        return pd.Timestamp.today().strftime("%Y-%m-%d")
    return END_DATE                        # 写死的具体日期

def get_raw(code, indicators, end=None, overlap_days=OVERLAP_DAYS, verbose=True):
    """带本地缓存的原始数据获取：增量刷新，返回完整历史 DataFrame(升序)。"""
    end = end or _default_end()
    cache = _load_raw_cache(code)
    # 缓存里可能缺少新加入的指标列 → 视为需要全量重拉
    need_full = (cache is None or cache.empty or
                 any(k not in cache.columns for k in indicators))
    if need_full:
        begin, mode = DATA_BEGIN, "全量首拉"
    else:
        last = cache.index.max()
        begin = (last - pd.Timedelta(days=overlap_days)).strftime("%Y-%m-%d")
        mode = f"增量(自 {begin})"
    new = _fetch_raw_block(code, indicators, begin, end)
    full = _combine_cache(None if need_full else cache, new)
    _save_raw_cache(code, full)
    if verbose:
        n_new = len(full) if cache is None else len(full) - len(cache)
        tail = full.index.max().date() if len(full) else "—"
        print(f"   · {code:<14}{mode:<16} 新增 {max(n_new,0):>3} 行, 共 {len(full):>5} 行 (至 {tail})")
    return full

def _build_data(idx_raw, mkt_raw):
    """把原始 DataFrame(升序) 转成 compute_temperature 所需的降序 dict。"""
    idx = pd.DatetimeIndex(sorted(idx_raw["close"].dropna().index, reverse=True))
    data = {"dates": idx}
    for k in IFIND_INDICATORS:
        data[k] = idx_raw[k].reindex(idx).to_numpy(dtype=float)
    for k in IFIND_MARKET_INDICATORS:
        data[k] = mkt_raw[k].reindex(idx).to_numpy(dtype=float)
    return data

def fetch_index(code, end=None, overlap_days=OVERLAP_DAYS, mkt_raw=None):
    """单指数取数(带缓存)。返回 compute_temperature 所需 data。"""
    if mkt_raw is None:
        mkt_raw = get_raw(MARKET_CODE, IFIND_MARKET_INDICATORS, end, overlap_days)
    idx_raw = get_raw(code, IFIND_INDICATORS, end, overlap_days)
    return _build_data(idx_raw, mkt_raw)

def run_batch(codes, end=None, overlap_days=OVERLAP_DAYS, history=False,
              user=None, pwd=None, out_path=None):
    """实盘批量(带本地增量缓存)：输入代码列表 → 输出温度汇总 Excel，返回最新温度 DataFrame。"""
    out_path = out_path or OUTPUT_PATH
    ifind_login(user, pwd)
    print("刷新数据（增量）：")
    mkt_raw = get_raw(MARKET_CODE, IFIND_MARKET_INDICATORS, end, overlap_days)  # 全市场只刷一次
    rows, hist = [], {}
    for code in codes:
        try:
            data = _build_data(get_raw(code, IFIND_INDICATORS, end, overlap_days), mkt_raw)
            df = compute_temperature(data, WEIGHTS, WIN)
            rows.append({"指数代码": code, **df.iloc[0].to_dict()})
            if history: hist[code] = df
            r0 = df.iloc[0]
            print(f"     → {code}  {r0['日期'].date()}  温度={r0['温度']:.4f}  "
                  f"乖离率={r0['乖离率%']:.2f}%({r0['乖离率分档']}/{r0['趋势方向']})")
        except Exception as e:
            rows.append({"指数代码": code, "温度": None, "错误": str(e)})
            print(f"     → [FAIL] {code}: {e}")
    latest = pd.DataFrame(rows)
    front = ["指数代码", "日期", "温度", "乖离率%", "乖离率分档", "趋势方向"]
    latest = latest[[c for c in front if c in latest] +
                    [c for c in latest.columns if c not in front]]
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        latest.to_excel(w, sheet_name="最新温度", index=False)
        for code, df in hist.items():
            df.to_excel(w, sheet_name=code[:31], index=False)
    print(f"\n已输出 → {out_path}（缓存目录: {CACHE_DIR}/）")
    return latest


# =============================================================================
#  ④ 离线演算（静态数据）—— 用模板自带缓存复算并打印每个细分指标读数
# =============================================================================
def load_static_from_xlsx(path=XLSX_PATH):
    """从模板缓存读取各底稿原始序列，构造 compute_temperature 所需 data。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    def col(sheet, c, r0=3):
        ws = wb[sheet]; out = {}
        for r in range(r0, ws.max_row + 1):
            dt = ws.cell(row=r, column=1).value
            v = ws.cell(row=r, column=c).value
            if dt is None: continue
            out[pd.Timestamp(dt)] = v if isinstance(v, (int, float)) else np.nan
        return pd.Series(out)

    series = dict(
        close=col("250日月跌幅均指滚动百分位", 2),
        amt=col("平均成交额", 2) * 1e8,        # 模板存"亿" → 还原为元
        turnover=col("日均换手率", 2),
        raise_num=col("上涨股票占比", 2),
        total_num=col("上涨股票占比", 4),
        limit_up=col("涨停家数和跌停家数", 2),
        limit_dn=col("涨停家数和跌停家数", 3),
        pe=col("pe", 2),
        pb=col("pb", 2),
        mkt_amt=col("平均成交额", 6) * 1e8,
        mkt_pb=col("pb", 5),
    )
    idx = set(series["close"].index)
    for k in ("amt", "turnover", "raise_num", "limit_up", "pe", "pb", "mkt_amt"):
        idx &= set(series[k].index)
    idx = pd.DatetimeIndex(sorted(idx, reverse=True))
    data = {"dates": idx}
    for k, s in series.items():
        data[k] = s.reindex(idx).to_numpy(dtype=float)
    e5 = wb["市场情绪温度计"]["E5"].value   # 模板缓存的温度，用于对照
    return data, e5


def demo_offline(path=XLSX_PATH):
    """静态数据演算 + 打印每个细分指标读数。"""
    data, e5 = load_static_from_xlsx(path)
    df = compute_temperature(data, WEIGHTS, WIN)
    latest = df.iloc[0]

    print("=" * 56)
    print(f"  指数: 000998.CSI（模板 F3）   最新交易日: {latest['日期'].date()}")
    print("=" * 56)
    print(f"{'细分指标':<14}{'读数':>10}{'权重':>8}")
    print("-" * 56)
    for name, w in WEIGHTS.items():
        v = latest[name]
        vs = "  N/A" if pd.isna(v) else f"{v:7.4f}"
        flag = "" if w else "   (不计权)"
        print(f"{name:<14}{vs:>10}{w:>8}{flag}")
    print("-" * 56)
    print(f"{'温度(本代码)':<14}{latest['温度']:>10.4f}")
    print(f"{'温度(模板E5)':<14}{e5:>10}")
    print(f"{'差异':<14}{abs(latest['温度'] - e5):>10.4f}  (仅模板四舍五入显示所致)")
    print("-" * 56)
    print("乖离率（减法版，独立指标，不并入温度）：")
    print(f"  乖离率   = {latest['乖离率%']:.2f}%")
    print(f"  分档     = {latest['乖离率分档']}   (>15过热 / 5~15强势 / -5~5偏弱 / <-5止损)")
    print(f"  趋势方向 = {latest['趋势方向']}   (近20日乖离率>0≥10日为'上行')")
    print("=" * 56)

    print("\n最近 6 个交易日温度 / 乖离率走势：")
    show = df[["日期", "温度", "乖离率%", "乖离率分档", "趋势方向"]].head(6).copy()
    show["日期"] = show["日期"].dt.date
    show["温度"] = show["温度"].round(4)
    print(show.to_string(index=False))
    return df


# =============================================================================
#  运行入口
#    · 实盘（办公电脑）：调 run_batch([...])，首次建缓存、之后每次只增量刷新
#    · 离线核对引擎  ：调 demo_offline()（需 BASE_DIR 下有模板 xlsx）
# =============================================================================
if __name__ == "__main__":
    # —— 实盘：填好上面「用户配置区」的 BASE_DIR 和 iFinD 账号后，用这行 ——
    run_batch(["000998.CSI", "399006.SZ", "000300.SH"], history=True)

    # —— 仅离线验证引擎（不连 iFinD）时，改用下面这行 ——
    # demo_offline()
