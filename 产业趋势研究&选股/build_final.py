import json, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
con = json.load(open('concepts.json', encoding='utf-8'))
res = json.load(open('results.json', encoding='utf-8'))
wb = openpyxl.Workbook(); ws = wb.active; ws.title='前五大重仓股'
hdr=['概念代码','概念名称','YTD涨跌幅']
for i in range(1,6): hdr+=[f'第{i}大重仓-名称',f'第{i}大重仓-代码',f'第{i}大重仓-流通市值(亿元)']
ws.append(hdr)
bold=Font(bold=True,color='FFFFFF'); fill=PatternFill('solid',fgColor='1F4E78')
center=Alignment(horizontal='center',vertical='center')
thin=Side(style='thin',color='D9D9D9'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
for c in ws[1]: c.font=bold; c.fill=fill; c.alignment=center; c.border=border
blank=[]
for c in con:
    code=c['code']; row=[code,c['name'],c.get('ytd')]; r=res.get(code)
    if r and r.get('holdings'):
        for h in r['holdings']: row+=[h['name'],h['code'],h['mktcap_yi']]
        while len(row)<18: row+=['','','']
    else: blank.append(code)
    ws.append(row)
ws.freeze_panes='A2'
ws.column_dimensions['A'].width=12; ws.column_dimensions['B'].width=20; ws.column_dimensions['C'].width=11
for i in range(4,19):
    ws.column_dimensions[get_column_letter(i)].width=13 if (i-4)%3==0 else (11 if (i-4)%3==1 else 16)
for r in range(2,ws.max_row+1):
    v=ws.cell(r,3).value
    if isinstance(v,(int,float)): ws.cell(r,3).number_format='0.00%'
    for cc in range(1,ws.max_column+1):
        ws.cell(r,cc).border=border; ws.cell(r,cc).alignment=Alignment(vertical='center')
ws3=wb.create_sheet('说明')
for n in ['数据来源：同花顺 iFinD（智能选股），数据日期 2026-06-18。',
 '口径：同花顺概念指数为等权指数，本表以“成分股中 A股流通市值(不含限售股)最大的前5只”作为前五大重仓股的近似。',
 '流通市值单位：亿元（= iFinD “a股市值(不含限售股)”）。',
 f'完成度：{len(res)}/{len(con)} 个概念已取数；其余 {len(blank)} 个按要求留空。']:
    ws3.append([n])
ws3.column_dimensions['A'].width=110
wb.save('同花顺A股概念指数_前五大重仓股.xlsx')
print('Saved. filled=%d/%d blank=%d' % (len(res),len(con),len(blank)))
