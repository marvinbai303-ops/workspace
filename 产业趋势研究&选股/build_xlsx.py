import json, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

con = json.load(open('concepts.json', encoding='utf-8'))
res = json.load(open('results.json', encoding='utf-8'))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '前五大重仓股'

hdr = ['概念代码','概念名称','YTD涨跌幅']
for i in range(1,6):
    hdr += [f'第{i}大重仓-名称', f'第{i}大重仓-代码', f'第{i}大重仓-流通市值(亿元)']
ws.append(hdr)

bold = Font(bold=True, color='FFFFFF')
fill = PatternFill('solid', fgColor='1F4E78')
center = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin,right=thin,top=thin,bottom=thin)
for c in ws[1]:
    c.font=bold; c.fill=fill; c.alignment=center; c.border=border

pending=[]
for c in con:
    code=c['code']; name=c['name']; ytd=c.get('ytd')
    row=[code, name, ytd]
    r=res.get(code)
    if r and r.get('holdings'):
        for h in r['holdings']:
            row += [h['name'], h['code'], h['mktcap_yi']]
        # pad if <5
        while len(row) < 3+5*3:
            row += ['','','']
    else:
        pending.append((code,name))
        row += ['待补充(数据源限流，需重试)']
    ws.append(row)

# formatting
ws.freeze_panes='A2'
ws.column_dimensions['A'].width=12
ws.column_dimensions['B'].width=20
ws.column_dimensions['C'].width=11
for i in range(4, 4+15):
    col=get_column_letter(i)
    ws.column_dimensions[col].width = 13 if (i-4)%3==0 else (11 if (i-4)%3==1 else 16)
# YTD as percent
for r in range(2, ws.max_row+1):
    v=ws.cell(r,3).value
    if isinstance(v,(int,float)):
        ws.cell(r,3).number_format='0.00%'
    for cc in range(1, ws.max_column+1):
        ws.cell(r,cc).border=border
        ws.cell(r,cc).alignment=Alignment(vertical='center')

# pending sheet
ws2=wb.create_sheet('待补充清单')
ws2.append(['概念代码','概念名称','状态'])
for c in ws2[1]:
    c.font=bold; c.fill=fill; c.alignment=center
for code,name in pending:
    ws2.append([code,name,'数据源限流未取到，待重试'])
ws2.column_dimensions['A'].width=12; ws2.column_dimensions['B'].width=22; ws2.column_dimensions['C'].width=24

# note
ws3=wb.create_sheet('说明')
notes=[
 '数据来源：同花顺 iFinD（智能选股），数据日期 2026-06-18',
 '口径：同花顺概念指数为等权指数，本表以“成分股中 A股流通市值(不含限售股)最大的前5只”作为前五大重仓股的近似。',
 '流通市值单位：亿元（= iFinD “a股市值(不含限售股)”）。',
 f'完成度：{len(res)}/{len(con)} 个概念已取数；其余见“待补充清单”，因数据源限流未取到，可稍后重试补全。',
 '个别概念说明：',
 '  886082 同花顺果指数 / 886093 华为数字能源：iFinD智能选股无法稳定解析为独立成分板块，待人工核对。',
 '  886073 铜缆高速连接：多次重试均未返回（板块存在但选股接口未解析）。',
]
for n in notes: ws3.append([n])
ws3.column_dimensions['A'].width=110

wb.save('同花顺A股概念指数_前五大重仓股.xlsx')
print('Saved. done=%d/%d, pending=%d' % (len(res), len(con), len(pending)))
