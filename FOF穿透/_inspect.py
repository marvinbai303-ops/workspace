import openpyxl, sys
out = open('_inspect_out.txt', 'w', encoding='utf-8')
def p(*a):
    print(*a)
    print(*a, file=out)
wb = openpyxl.load_workbook('中信三级行业指数.xlsx', read_only=True, data_only=True)
p('Sheets:', wb.sheetnames)
for ws in wb.worksheets:
    p('===', ws.title, 'dims:', ws.max_row, 'x', ws.max_column)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 20:
            p('... (more rows)')
            break
        p(i, list(row))
out.close()
print('DONE')
