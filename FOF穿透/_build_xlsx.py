"""Build CR5 xlsx (wide format): one row per index, top-5 expanded horizontally."""
import json, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

cr5 = json.load(open("/Users/yangguang/Documents/Claude/Projects/FOF穿透/_cr5.json", encoding="utf-8"))
indices = json.load(open("/Users/yangguang/Documents/Claude/Projects/FOF穿透/_indices.json", encoding="utf-8"))

by_idx = {}
for r in cr5:
    by_idx.setdefault(r["index_code"], []).append(r)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "中信三级行业CR5"

hdr = ["指数代码", "指数名称", "xlsx成分数", "三级行业成分数"]
for i in range(1, 6):
    hdr += [f"CR{i}代码", f"CR{i}名称", f"CR{i}自由流通市值(亿)", f"CR{i}权重(%)"]
hdr.append("CR5合计权重(%)")
ws.append(hdr)

header_fill = PatternFill("solid", fgColor="DDEBF7")
group_fills = ["FFF2CC", "E2EFDA", "FCE4D6", "DDEBF7", "EAD1DC"]  # 5 group colors
for c in range(1, len(hdr) + 1):
    cell = ws.cell(1, c)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if c <= 4:
        cell.fill = header_fill
    elif c == len(hdr):
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
    else:
        grp = (c - 5) // 4
        cell.fill = PatternFill("solid", fgColor=group_fills[grp])

for it in indices:
    code = it["code"]
    rows = by_idx.get(code, [])
    rows = sorted([r for r in rows if r.get("rank")], key=lambda x: x["rank"])
    n_ind = rows[0]["n_industry"] if rows else ""
    line = [code, it["name"], it["n"], n_ind]
    total_w = 0
    for i in range(5):
        if i < len(rows):
            r = rows[i]
            w = r.get("weight_pct") or 0
            total_w += w
            line += [r["stock_code"], r["stock_name"], r["mcap_yi"], w]
        else:
            line += ["", "", "", ""]
    line.append(round(total_w, 2))
    ws.append(line)

# column widths
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 11
ws.column_dimensions["D"].width = 13
for i in range(5):
    base = 5 + i * 4
    ws.column_dimensions[openpyxl.utils.get_column_letter(base)].width = 11
    ws.column_dimensions[openpyxl.utils.get_column_letter(base + 1)].width = 14
    ws.column_dimensions[openpyxl.utils.get_column_letter(base + 2)].width = 17
    ws.column_dimensions[openpyxl.utils.get_column_letter(base + 3)].width = 11
ws.column_dimensions[openpyxl.utils.get_column_letter(len(hdr))].width = 15
ws.freeze_panes = "E2"

out = "/Users/yangguang/Documents/Claude/Projects/FOF穿透/中信三级行业指数_CR5.xlsx"
wb.save(out)
print("written:", out)
print("rows:", ws.max_row - 1)
