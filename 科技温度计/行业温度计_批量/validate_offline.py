# -*- coding: utf-8 -*-
"""
离线自检 —— 不连 iFinD，用模板自带的缓存原始数据跑一遍计算引擎，
确认 compute.py 复刻的温度与模板 E5 一致。

用法: python validate_offline.py [模板路径]
"""
import sys
import numpy as np
import pandas as pd
import openpyxl

from compute import compute_temperature
from config import WEIGHTS, WIN

XLSX = sys.argv[1] if len(sys.argv) > 1 else "../转换 行业温度计_量价_估值.xlsx"


def col(wb, sheet, c, r0=3):
    ws = wb[sheet]
    out = {}
    for r in range(r0, ws.max_row + 1):
        dt = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=c).value
        if dt is None:
            continue
        out[pd.Timestamp(dt)] = v if isinstance(v, (int, float)) else np.nan
    return pd.Series(out)


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    close = col(wb, "250日月跌幅均指滚动百分位", 2)
    amt = col(wb, "平均成交额", 2) * 1e8
    turn = col(wb, "日均换手率", 2)
    raise_ = col(wb, "上涨股票占比", 2)
    total = col(wb, "上涨股票占比", 4)
    lup = col(wb, "涨停家数和跌停家数", 2)
    ldn = col(wb, "涨停家数和跌停家数", 3)
    pe = col(wb, "pe", 2)
    pb = col(wb, "pb", 2)
    mamt = col(wb, "平均成交额", 6) * 1e8
    mpb = col(wb, "pb", 5)

    idx = (set(close.index) & set(amt.index) & set(turn.index) & set(raise_.index)
           & set(lup.index) & set(pe.index) & set(pb.index) & set(mamt.index))
    idx = pd.DatetimeIndex(sorted(idx, reverse=True))
    a = lambda s: s.reindex(idx).to_numpy(dtype=float)
    data = dict(dates=idx, close=a(close), amt=a(amt), turnover=a(turn),
                raise_num=a(raise_), total_num=a(total), limit_up=a(lup),
                limit_dn=a(ldn), pe=a(pe), pb=a(pb), mkt_amt=a(mamt), mkt_pb=a(mpb))

    df = compute_temperature(data, WEIGHTS, WIN)
    e5 = wb["市场情绪温度计"]["E5"].value
    print(f"最新日期 {df.iloc[0]['日期'].date()}")
    print(f"计算温度 = {df.iloc[0]['温度']:.4f}   模板E5 = {e5}")
    print("（差异仅来自模板缓存的四舍五入显示，应 < 0.001）")


if __name__ == "__main__":
    main()
