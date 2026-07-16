#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从已刷新并保存的 Excel 模板读取组合级结果，生成样板间组合日报图片。

关键口径：
  - 组合净值、日收益直接取 Excel「实盘组合净值」最新日期行；
  - 累计收益用 Excel 净值 - 1；
  - 持仓明细只用于展示，由 Excel「分期持仓明细」最新生效权重和基金净值表派生。
"""
import argparse
import datetime as dt
import json
import os

from openpyxl import load_workbook

import daily_monitor as dm
import gen_images

PORTFOLIO_ORDER = ['产业趋势', '产业趋势基石', '产业趋势2号', '10-90', '10-90基石', '30-70']
MONEY_CODES = {'000917.OF'}


def as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return dt.datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def ymd(value):
    return value.strftime("%Y%m%d")


def norm_code(code):
    if code is None:
        return None
    code = str(code).strip()
    if not code:
        return None
    return code if "." in code else f"{code}.OF"


def load_meta(wb):
    ws = wb['实盘持仓配置']
    meta = {}
    for row in range(4, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not isinstance(label, str) or '权重' not in label:
            continue
        name = label.replace('权重', '').strip()
        start = as_date(ws.cell(row, 27).value)
        meta[name] = {
            'fee': float(ws.cell(row, 26).value or 0),
            'start': ymd(start) if start else '',
        }
    return meta


def dated_rows(ws, date_col=1, min_row=1, max_row=None):
    rows = []
    for row in range(min_row, (max_row or ws.max_row) + 1):
        d = as_date(ws.cell(row, date_col).value)
        if d:
            rows.append((d, row))
    return rows


def latest_portfolio_row(wb):
    ws = wb['实盘组合净值']
    candidates = []
    for d, row in dated_rows(ws, min_row=5):
        has_nav = any(ws.cell(row, col).value not in (None, '') for col in (2, 5, 8, 11, 14, 17))
        if has_nav:
            candidates.append((d, row))
    if not candidates:
        raise RuntimeError('实盘组合净值没有找到有效日期行')
    return max(candidates, key=lambda x: x[0])


def read_fund_navs(wb):
    ws = wb['实盘净值数据(iFinD公式)']
    codes = {}
    names = {}
    for col in range(2, ws.max_column + 1):
        code = norm_code(ws.cell(3, col).value)
        if not code:
            continue
        codes[col] = code
        names[code] = ws.cell(4, col).value or code

    navs = {}
    for d, row in dated_rows(ws, min_row=5):
        day = {}
        for col, code in codes.items():
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)):
                day[code] = float(value)
        if day:
            navs[d] = day
    return names, navs


def read_active_weights(wb, portfolio, asof):
    ws = wb['分期持仓明细']
    names = {}
    codes = {}
    for col in range(2, 26):
        code = norm_code(ws.cell(3, col).value)
        if not code:
            continue
        codes[col] = code
        names[code] = ws.cell(2, col).value or code

    candidates = []
    for row in range(5, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not isinstance(label, str) or not label.startswith(f'{portfolio} ·'):
            continue
        eff = as_date(ws.cell(row, 27).value)
        if eff and eff <= asof:
            candidates.append((eff, row))
    if not candidates:
        return {}, names, None

    eff, row = max(candidates, key=lambda x: x[0])
    weights = {}
    for col, code in codes.items():
        value = ws.cell(row, col).value
        if isinstance(value, (int, float)) and abs(value) > 1e-10:
            weights[code] = float(value)
    return weights, names, eff


def build_holdings(wb, portfolio, asof, fund_names, navs):
    weights, weight_names, effective_date = read_active_weights(wb, portfolio, asof)
    if not weights:
        return []

    nav_dates = sorted(d for d in navs if d <= asof)
    if not nav_dates:
        return []
    latest = nav_dates[-1]
    prev = nav_dates[-2] if len(nav_dates) >= 2 else latest
    base_dates = [d for d in nav_dates if effective_date is None or d <= effective_date]
    base = base_dates[-1] if base_dates else nav_dates[0]

    holdings = []
    for code, weight in weights.items():
        now = navs.get(latest, {}).get(code)
        before = navs.get(prev, {}).get(code)
        start = navs.get(base, {}).get(code)
        day_ret = (now / before - 1.0) if now and before else 0.0
        cum_ret = (now / start - 1.0) if now and start else 0.0
        holdings.append({
            'code': code,
            'name': fund_names.get(code) or weight_names.get(code) or code,
            'w': weight,
            'day': day_ret,
            'cum': cum_ret,
            'cd': weight * day_ret,
            'cc': weight * cum_ret,
            'money': code in MONEY_CODES,
        })
    holdings.sort(key=lambda h: (h['money'], -h['day']))
    return holdings


def build_report(excel_path):
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    meta = load_meta(wb)
    asof, row = latest_portfolio_row(wb)
    fund_names, navs = read_fund_navs(wb)
    ws = wb['实盘组合净值']

    result = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(3, col).value
        if not isinstance(label, str) or not label.startswith('组合 '):
            continue
        name = label.replace('组合 ', '').strip()
        nav = ws.cell(row, col).value
        day_ret = ws.cell(row, col + 1).value
        if not isinstance(nav, (int, float)):
            continue
        m = meta.get(name, {})
        result[name] = {
            'asof': ymd(asof),
            'start': m.get('start', ''),
            'fee': float(m.get('fee', 0)),
            'nav': float(nav),
            'day_ret': float(day_ret or 0),
            'cum': float(nav) - 1.0,
            'holdings': build_holdings(wb, name, asof, fund_names, navs),
            'source': 'excel',
        }
    if not result:
        raise RuntimeError('实盘组合净值没有读取到任何组合结果')
    return result, ymd(asof)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--excel', required=True, help='已刷新并保存的 Excel 模板')
    ap.add_argument('--outdir', default='reports')
    a = ap.parse_args()

    os.makedirs(a.outdir, exist_ok=True)
    res, asof = build_report(a.excel)

    print(f'=== Excel 数据日期 {asof} ===')
    print(f'{"组合":<12}{"模拟净值":>10}{"今日涨跌":>10}{"累计涨幅":>10}')
    for p in PORTFOLIO_ORDER:
        if p not in res:
            continue
        r = res[p]
        print(f'{p:<12}{r["nav"]:>10.4f}{dm.pct(r["day_ret"]):>11}{dm.pct(r["cum"]):>11}')

    outs = [gen_images.gen_overview(res, asof, a.outdir)]
    for key in gen_images.ORDER:
        outs.append(gen_images.gen_strategy(key, res, asof, a.outdir))
    for path in outs:
        if path:
            print('PNG ->', path)

    audit_path = os.path.join(a.outdir, 'daily_report_from_excel.json')
    with open(audit_path, 'w', encoding='utf-8') as f:
        json.dump({'asof': asof, 'source_excel': os.path.abspath(a.excel), 'portfolios': res}, f, ensure_ascii=False, indent=2)
    print('JSON ->', audit_path)


if __name__ == '__main__':
    main()
