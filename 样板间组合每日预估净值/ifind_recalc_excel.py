#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用 iFinD 最新底层基金复权单位净值，按 Excel 模板公式口径重算组合预估净值。

脚本只写回 xlsx 内部 XML 的日期与公式缓存值，不用 openpyxl 另存，避免破坏
模板中既有的 iFinD/Excel 公式。
"""
import argparse
import datetime as dt
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

import excel_daily_report as report

NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
ET.register_namespace('', NS['main'])
ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
ET.register_namespace('x15', 'http://schemas.microsoft.com/office/spreadsheetml/2010/11/main')
ET.register_namespace('x15ac', 'http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac')

FUND_SHEET = '实盘净值数据(iFinD公式)'
COMBO_SHEET = '实盘组合净值'
WEIGHT_SHEET = '分期持仓明细'
CONFIG_SHEET = '实盘持仓配置'
PORTFOLIO_ORDER = ['产业趋势', '产业趋势基石', '产业趋势2号', '10-90', '10-90基石', '30-70']
MONEY_CODES = {'000917.OF'}
EXPECTED_FUND_COUNT = 24


def as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        return dt.date(1899, 12, 30) + dt.timedelta(days=int(value))
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%Y%m%d', '%Y-%m-%d', '%Y/%m/%d'):
            try:
                return dt.datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def ymd(value):
    return value.strftime('%Y%m%d')


def parse_date(text):
    return dt.datetime.strptime(text, '%Y%m%d').date()


def norm_code(code):
    if code is None:
        return None
    code = str(code).strip()
    if not code:
        return None
    return code if '.' in code else f'{code}.OF'


def excel_serial(day):
    return (day - dt.date(1899, 12, 30)).days


def col_name(index):
    out = ''
    while index:
        index, rem = divmod(index - 1, 26)
        out = chr(65 + rem) + out
    return out


def num_text(value):
    if value is None:
        return ''
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError(f'invalid numeric value: {value}')
        return f'{value:.15g}'
    return str(value)


def atomic_write_json(path, payload):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + '.tmp')
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    os.replace(tmp, path)


def extract_answer_text(result):
    texts = []
    for batch in result.get('batches', []):
        res = batch.get('result') or {}
        data = res.get('data') or {}
        content = ((data.get('result') or {}).get('content') or [])
        for item in content:
            text = item.get('text') if isinstance(item, dict) else None
            if not text:
                continue
            try:
                nested = json.loads(text)
                answer = ((nested.get('data') or {}).get('answer1')
                          or nested.get('answer')
                          or text)
                texts.append(str(answer))
            except json.JSONDecodeError:
                texts.append(text)
    return '\n'.join(texts)


def parse_navs_from_text(text, target_ymd=None):
    navs = {}
    dates = {}
    header = None
    code_re = re.compile(r'\b(\d{6}(?:\.OF)?)\b')
    date_re = re.compile(r'\b(20\d{6})\b')
    number_re = re.compile(r'(?<![A-Za-z0-9.])-?\d+(?:\.\d+)?')

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if '|' in line:
            cells = [c.strip() for c in line.strip('|').split('|')]
            if any('证券代码' in c or '基金代码' in c for c in cells):
                header = cells
                continue
            if set(cells) <= {'---', ':---', '---:', ':---:'}:
                continue
            joined = '|'.join(cells)
            m = code_re.search(joined)
            if not m:
                continue
            code = norm_code(m.group(1))
            nav_idx = None
            date_idx = None
            if header and len(header) == len(cells):
                for i, h in enumerate(header):
                    if '净值' in h and ('复权' in h or '单位' in h):
                        nav_idx = i
                        break
                for i, h in enumerate(header):
                    if '日期' in h:
                        date_idx = i
                        break
            candidates = []
            if nav_idx is not None:
                candidates = number_re.findall(cells[nav_idx])
            if not candidates:
                candidates = [
                    n for n in number_re.findall(joined)
                    if n not in (code[:6],)
                    and not date_re.fullmatch(n)
                ]
            if candidates:
                row_date = None
                if date_idx is not None:
                    dm = date_re.search(cells[date_idx])
                    row_date = dm.group(1) if dm else None
                if row_date is None:
                    dm = date_re.search(joined)
                    row_date = dm.group(1) if dm else None
                if target_ymd and row_date and row_date == target_ymd:
                    navs[code] = float(candidates[-1])
                    dates[code] = row_date
                elif code not in navs:
                    navs[code] = float(candidates[-1])
                    if row_date:
                        dates[code] = row_date
            continue

        m = code_re.search(line)
        if not m:
            continue
        code = norm_code(m.group(1))
        candidates = [n for n in number_re.findall(line) if n != code[:6]]
        if candidates:
            row_date = None
            dm = date_re.search(line)
            if dm:
                row_date = dm.group(1)
            if target_ymd and row_date and row_date == target_ymd:
                navs[code] = float(candidates[-1])
                dates[code] = row_date
            elif code not in navs:
                navs[code] = float(candidates[-1])
                if row_date:
                    dates[code] = row_date
    return navs, dates


def fetch_ifind_navs(codes, target_date, outdir, batch_size=6, delay_ms=1300):
    payload = {
        'target_date': ymd(target_date),
        'codes': codes,
        'batch_size': batch_size,
        'delay_ms': delay_ms,
    }
    cmd = ['node', 'ifind_nav_query.js', json.dumps(payload, ensure_ascii=False)]
    proc = subprocess.run(cmd, text=True, capture_output=True)
    raw_path = Path(outdir) / f'ifind_raw_{ymd(target_date)}.json'
    if proc.stdout.strip():
        raw_path.parent.mkdir(parents=True, exist_ok=True)
        raw_path.write_text(proc.stdout, encoding='utf-8')
    if proc.returncode != 0:
        raise RuntimeError(f'iFinD 查询失败：{proc.stderr.strip() or proc.stdout.strip()}')
    result = json.loads(proc.stdout)
    text = extract_answer_text(result)
    navs, dates = parse_navs_from_text(text, ymd(target_date))
    return navs, dates, result


def load_nav_json(path):
    payload = json.loads(Path(path).read_text(encoding='utf-8'))
    if 'navs' in payload:
        payload = payload['navs']
    return {norm_code(k): float(v) for k, v in payload.items()}


def read_fund_codes(wb):
    ws = wb[FUND_SHEET]
    codes = []
    names = {}
    cols = {}
    for col in range(2, ws.max_column + 1):
        code = norm_code(ws.cell(3, col).value)
        if not code:
            continue
        codes.append(code)
        cols[code] = col
        names[code] = ws.cell(4, col).value or code
    return codes, cols, names


def dated_rows(ws, min_row=5):
    rows = []
    for row in range(min_row, ws.max_row + 1):
        day = as_date(ws.cell(row, 1).value)
        if day:
            rows.append((day, row))
    return rows


def target_or_blank_row(ws, target_date, min_row=5):
    first_blank = None
    for row in range(min_row, ws.max_row + 1):
        day = as_date(ws.cell(row, 1).value)
        if day == target_date:
            return row
        if day is None and first_blank is None:
            first_blank = row
    if first_blank is None:
        raise RuntimeError(f'{ws.title} 没有可写入的空日期行')
    return first_blank


def read_fund_history(wb, codes):
    ws = wb[FUND_SHEET]
    history = {}
    for day, row in dated_rows(ws):
        values = {}
        for col in range(2, ws.max_column + 1):
            code = norm_code(ws.cell(3, col).value)
            if code not in codes:
                continue
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)):
                values[code] = float(value)
        if values:
            history[day] = values
    return history


def read_config(wb):
    ws = wb[CONFIG_SHEET]
    config = {}
    for row in range(4, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not isinstance(label, str) or '权重' not in label:
            continue
        name = label.replace('权重', '').strip()
        config[name] = {
            'fee': float(ws.cell(row, 26).value or 0),
            'start': as_date(ws.cell(row, 27).value),
        }
    return config


def active_weight_row(wb, portfolio, asof):
    ws = wb[WEIGHT_SHEET]
    candidates = []
    for row in range(5, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not isinstance(label, str) or not label.startswith(f'{portfolio} ·'):
            continue
        eff = as_date(ws.cell(row, 27).value)
        if eff and eff <= asof:
            candidates.append((eff, row))
    if not candidates:
        raise RuntimeError(f'未找到 {portfolio} 在 {ymd(asof)} 前生效的持仓行')
    eff, row = max(candidates, key=lambda item: item[0])
    weights = {}
    for col in range(2, 26):
        code = norm_code(ws.cell(3, col).value)
        value = ws.cell(row, col).value
        if code and isinstance(value, (int, float)) and abs(value) > 1e-12:
            weights[code] = float(value)
    factor = float(ws.cell(row, 28).value or 1.0)
    return eff, row, weights, factor


def read_portfolio_history(wb):
    ws = wb[COMBO_SHEET]
    cols = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(3, col).value
        if isinstance(label, str) and label.startswith('组合 '):
            cols[label.replace('组合 ', '').strip()] = col

    history = {name: {} for name in cols}
    for day, row in dated_rows(ws):
        for name, col in cols.items():
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)):
                history[name][day] = float(value)
    return cols, history


def calculate_portfolios(wb, target_date, fund_history, current_navs):
    config = read_config(wb)
    portfolio_cols, portfolio_history = read_portfolio_history(wb)
    fund_history = dict(fund_history)
    fund_history[target_date] = current_navs
    results = {}

    for name in PORTFOLIO_ORDER:
        if name not in portfolio_cols:
            raise RuntimeError(f'实盘组合净值缺少组合列：{name}')
        if name not in config or not config[name].get('start'):
            raise RuntimeError(f'实盘持仓配置缺少 {name} 的费率或起始日')
        eff, weight_row, weights, factor = active_weight_row(wb, name, target_date)
        base_navs = fund_history.get(eff)
        if not base_navs:
            raise RuntimeError(f'{name} 的生效日 {ymd(eff)} 没有底层基金基准净值')

        portfolio_nav = 0.0
        missing = []
        for code, weight in weights.items():
            now = current_navs.get(code)
            base = base_navs.get(code)
            if not now or not base:
                missing.append(code)
                continue
            portfolio_nav += weight * now / base
        if missing:
            raise RuntimeError(f'{name} 缺少底层基金净值：{",".join(missing)}')

        start = config[name]['start']
        fee = config[name]['fee']
        nav = factor * portfolio_nav * ((1.0 - fee) ** ((target_date - start).days / 365.0))
        hist = {d: v for d, v in portfolio_history[name].items() if d < target_date}
        prev_day = max(hist) if hist else None
        day_ret = nav / hist[prev_day] - 1.0 if prev_day else None
        max_nav = max(list(hist.values()) + [nav])
        drawdown = nav / max_nav - 1.0
        results[name] = {
            'nav': nav,
            'day_ret': day_ret,
            'drawdown': drawdown,
            'effective_date': ymd(eff),
            'weight_row': weight_row,
            'factor': factor,
        }
    return results


def workbook_sheet_paths(xlsx_path):
    with zipfile.ZipFile(xlsx_path) as zin:
        workbook = ET.fromstring(zin.read('xl/workbook.xml'))
        rels = ET.fromstring(zin.read('xl/_rels/workbook.xml.rels'))

    rel_map = {}
    for rel in rels:
        rid = rel.attrib.get('Id')
        target = rel.attrib.get('Target')
        if rid and target and target.startswith('worksheets/'):
            rel_map[rid] = f'xl/{target}'

    paths = {}
    rel_ns = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
    for sheet in workbook.findall('.//main:sheet', NS):
        name = sheet.attrib.get('name')
        rid = sheet.attrib.get(rel_ns)
        if name and rid in rel_map:
            paths[name] = rel_map[rid]
    return paths


def find_row(root, row_num):
    sheet_data = root.find('main:sheetData', NS)
    for row in sheet_data.findall('main:row', NS):
        if row.attrib.get('r') == str(row_num):
            return row
    raise RuntimeError(f'XML 中找不到 row {row_num}')


def find_cell(row, ref):
    for cell in row.findall('main:c', NS):
        if cell.attrib.get('r') == ref:
            return cell
    raise RuntimeError(f'XML 中找不到 cell {ref}')


def set_cell_value(cell, value):
    cell.attrib.pop('t', None)
    v = cell.find('main:v', NS)
    if v is None:
        v = ET.SubElement(cell, f'{{{NS["main"]}}}v')
    v.text = num_text(value)


def patch_sheet(xml_bytes, updates):
    root = ET.fromstring(xml_bytes)
    for row_num, values in updates.items():
        row = find_row(root, row_num)
        for col, value in values.items():
            ref = f'{col_name(col)}{row_num}'
            set_cell_value(find_cell(row, ref), value)
    return ET.tostring(root, encoding='utf-8', xml_declaration=True)


def patch_workbook(xlsx_path, fund_row, combo_row, target_date, fund_cols, navs, portfolio_cols, results):
    paths = workbook_sheet_paths(xlsx_path)
    fund_path = paths[FUND_SHEET]
    combo_path = paths[COMBO_SHEET]
    fund_updates = {fund_row: {1: excel_serial(target_date)}}
    for code, col in fund_cols.items():
        fund_updates[fund_row][col] = navs[code]

    combo_updates = {combo_row: {1: excel_serial(target_date)}}
    for name, col in portfolio_cols.items():
        item = results.get(name)
        if not item:
            continue
        combo_updates[combo_row][col] = item['nav']
        combo_updates[combo_row][col + 1] = item['day_ret'] if item['day_ret'] is not None else ''
        combo_updates[combo_row][col + 2] = item['drawdown']

    xlsx_path = Path(xlsx_path)
    fd, tmp_name = tempfile.mkstemp(suffix='.xlsx', dir=str(xlsx_path.parent))
    os.close(fd)
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin, zipfile.ZipFile(tmp_name, 'w', zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename == fund_path:
                    data = patch_sheet(data, fund_updates)
                elif info.filename == combo_path:
                    data = patch_sheet(data, combo_updates)
                zout.writestr(info, data)
        os.replace(tmp_name, xlsx_path)
    except Exception:
        if os.path.exists(tmp_name):
            os.unlink(tmp_name)
        raise


def compare_stale(navs, fund_history, target_date):
    previous_days = [d for d in fund_history if d < target_date]
    if not previous_days:
        return []
    prev = fund_history[max(previous_days)]
    stale = []
    for code, value in navs.items():
        if code in MONEY_CODES:
            continue
        old = prev.get(code)
        if old is not None and abs(old - value) <= 1e-12:
            stale.append(code)
    return stale


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--excel', required=True)
    ap.add_argument('--target-date', required=True, help='YYYYMMDD')
    ap.add_argument('--outdir', default='reports/run_logs')
    ap.add_argument('--nav-json', default=None, help='可选：跳过 iFinD 查询，直接读取 code->nav JSON')
    ap.add_argument('--stale-equal-limit', type=int, default=3)
    ap.add_argument('--batch-size', type=int, default=6)
    ap.add_argument('--delay-ms', type=int, default=1300)
    ap.add_argument('--backup', action='store_true')
    args = ap.parse_args()

    target_date = parse_date(args.target_date)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    if args.backup:
        backup = Path(args.excel).with_suffix(f'.before_ifind_recalc_{args.target_date}.xlsx')
        if not backup.exists():
            shutil.copy2(args.excel, backup)

    wb = load_workbook(args.excel, data_only=True, read_only=True)
    codes, fund_cols, fund_names = read_fund_codes(wb)
    fund_history = read_fund_history(wb, set(codes))
    fund_ws = wb[FUND_SHEET]
    combo_ws = wb[COMBO_SHEET]
    fund_row = target_or_blank_row(fund_ws, target_date)
    combo_row = target_or_blank_row(combo_ws, target_date)

    if args.nav_json:
        navs = load_nav_json(args.nav_json)
        parsed_dates = {}
        raw_result = {'source': args.nav_json}
    else:
        navs, parsed_dates, raw_result = fetch_ifind_navs(codes, target_date, outdir, args.batch_size, args.delay_ms)

    missing = [code for code in codes if code not in navs]
    if missing:
        payload = {
            'status': 'FAILED',
            'target_date': args.target_date,
            'reason': f'iFinD 最新复权单位净值未齐全：{len(codes) - len(missing)}/{len(codes)}',
            'missing_codes': missing,
            'parsed_navs': navs,
            'parsed_dates': parsed_dates,
        }
        atomic_write_json(outdir / f'ifind_recalc_{args.target_date}.json', payload)
        raise RuntimeError(payload['reason'])

    wrong_date = [
        code for code in codes
        if parsed_dates.get(code) and parsed_dates[code] != args.target_date
    ]
    if wrong_date:
        payload = {
            'status': 'FAILED',
            'target_date': args.target_date,
            'reason': f'{len(wrong_date)} 只基金 iFinD 明确返回的净值日期不是目标日期',
            'wrong_date_codes': {
                code: parsed_dates.get(code)
                for code in wrong_date
            },
            'parsed_navs': navs,
            'parsed_dates': parsed_dates,
        }
        atomic_write_json(outdir / f'ifind_recalc_{args.target_date}.json', payload)
        raise RuntimeError(payload['reason'])

    navs = {code: navs[code] for code in codes}
    stale = compare_stale(navs, fund_history, target_date)
    if len(stale) >= args.stale_equal_limit:
        payload = {
            'status': 'FAILED',
            'target_date': args.target_date,
            'reason': f'{len(stale)} 只非货币基金最新净值与上一交易日完全相同，疑似 iFinD 最新值未完全刷新',
            'stale_equal_codes': stale,
            'parsed_navs': navs,
            'parsed_dates': parsed_dates,
        }
        atomic_write_json(outdir / f'ifind_recalc_{args.target_date}.json', payload)
        raise RuntimeError(payload['reason'])

    results = calculate_portfolios(wb, target_date, fund_history, navs)
    portfolio_cols, _ = read_portfolio_history(wb)
    patch_workbook(args.excel, fund_row, combo_row, target_date, fund_cols, navs, portfolio_cols, results)

    payload = {
        'status': 'SUCCESS',
        'target_date': args.target_date,
        'fund_row': fund_row,
        'combo_row': combo_row,
        'fund_count': len(navs),
        'stale_equal_codes': stale,
        'fund_navs': {
            code: {'name': fund_names.get(code, code), 'nav': navs[code], 'parsed_date': parsed_dates.get(code)}
            for code in codes
        },
        'portfolios': results,
        'raw_source': raw_result.get('source') if isinstance(raw_result, dict) else None,
    }
    atomic_write_json(outdir / f'ifind_recalc_{args.target_date}.json', payload)
    print(f'iFinD 重算写回完成：{args.target_date}，底层基金 {len(navs)} 只')
    for name in PORTFOLIO_ORDER:
        item = results[name]
        print(f'{name}: 净值 {item["nav"]:.4f}, 今日 {item["day_ret"] or 0:.2%}, 回撤 {item["drawdown"]:.2%}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
