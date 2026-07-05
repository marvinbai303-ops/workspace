# -*- coding: utf-8 -*-
"""Build slides 18 (<-16) and 19 (<-17) for 进取组合, using net-value series.
Faithful copy of slide16/17 styled shapes; only data swapped. No pyramid on 18."""
import openpyxl, math, datetime, shutil, copy, statistics
from collections import OrderedDict
from lxml import etree

U = "unpacked"
A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
P = "{http://schemas.openxmlformats.org/presentationml/2006/main}"
C = "{http://schemas.openxmlformats.org/drawingml/2006/chart}"
R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
NS = {'a':A[1:-1],'p':P[1:-1],'c':C[1:-1],'r':R[1:-1]}
parser = etree.XMLParser(remove_blank_text=True)

# ---------------- 1. Load net-value data ----------------
wb = openpyxl.load_workbook("进取组合回测净值序列.xlsx", data_only=True)
ws = wb["Sheet1"]
data = []
for r in range(2, ws.max_row+1):
    d = ws.cell(r,1).value; v = ws.cell(r,2).value
    if d is None or v is None: continue
    if isinstance(d, str): d = datetime.datetime.fromisoformat(d)
    data.append((d.date() if isinstance(d, datetime.datetime) else d, float(v)))
dates = [d for d,_ in data]; nav = [v for _,v in data]
assert nav[0] == 1.0, nav[0]

# ---------------- 2. Compute metrics ----------------
T = (dates[-1]-dates[0]).days/365.25
cagr = (nav[-1]/nav[0])**(1/T)-1
rets = [nav[i]/nav[i-1]-1 for i in range(1,len(nav))]
vol = statistics.pstdev(rets)*math.sqrt(252)
peak=nav[0]; mdd=0.0
for v in nav:
    peak=max(peak,v); mdd=min(mdd, v/peak-1)
sharpe = cagr/vol

# monthly last NAV
monthly = OrderedDict()
for d,v in data: monthly[(d.year,d.month)] = v
mret = OrderedDict(); prev = nav[0]
for k in monthly:
    mret[k] = monthly[k]/prev - 1; prev = monthly[k]
# drop the inception base month (2019-12) from "complete months"
inception_key = (dates[0].year, dates[0].month)  # (2019,12)

years = [y for y in sorted(set(y for y,_ in monthly)) if y >= 2020]
# annual return / winrate / intra-year maxdd
prev_ye = nav[0]
annual = OrderedDict()
for y in years:
    yv = [(d,v) for d,v in data if d.year==y]
    ye = yv[-1][1]
    aret = ye/prev_ye-1; prev_ye = ye
    ms = [mret[(y,m)] for m in range(1,13) if (y,m) in mret]
    wr = sum(1 for x in ms if x>0)/len(ms)
    seg=[v for d,v in yv]; pk=seg[0]; dd=0.0
    for v in seg:
        pk=max(pk,v); dd=min(dd, v/pk-1)
    annual[y] = dict(ret=aret, wr=wr, dd=dd, n=len(ms))

full_years = [y for y in years if annual[y]['n']==12]
arith_avg_annual = statistics.mean(annual[y]['ret'] for y in full_years)  # 2020-2025
cum = nav[-1]/nav[0]-1
allm = [mret[k] for k in mret if k != inception_key]
overall_wr = sum(1 for x in allm if x>0)/len(allm)
pos_cnt = sum(1 for x in allm if x>0)
mean_m = statistics.mean(allm)
pos_years = sum(1 for y in full_years if annual[y]['ret']>0)
best_y = max(full_years, key=lambda y: annual[y]['ret'])
worst_y= min(full_years, key=lambda y: annual[y]['ret'])

def pct(x, dp=2): return f"{x*100:.{dp}f}%"
print("CAGR",pct(cagr),"arithAvgAnnual",pct(arith_avg_annual),"vol",pct(vol),
      "mdd",pct(mdd),"sharpe",f"{sharpe:.2f}","cum",pct(cum),
      "overall_wr",pct(overall_wr),"pos_cnt",pos_cnt,"nmonths",len(allm),
      "mean_m",pct(mean_m),"pos_years",pos_years,
      "best",best_y,pct(annual[best_y]['ret']),"worst",worst_y,pct(annual[worst_y]['ret']))

# ---------------- helpers ----------------
def first_cnvpr(sh):
    return sh.find('.//'+P+'cNvPr')

def set_para_text(p, text, template_rpr=None):
    """Collapse paragraph runs to a single run carrying first run's rPr; set text."""
    runs = p.findall(A+'r')
    if runs:
        first = runs[0]
        for r in runs[1:]: p.remove(r)
        ts = first.findall(A+'t')
        for t in ts[1:]: first.remove(t)
        if ts: ts[0].text = text
        else:
            t = etree.SubElement(first, A+'t'); t.text = text
    else:
        # build a run from template rPr
        r = etree.SubElement(p, A+'r')
        if template_rpr is not None:
            r.append(copy.deepcopy(template_rpr))
        t = etree.SubElement(r, A+'t'); t.text = text

def get_table(frame):
    return frame.find('.//'+A+'tbl')

def set_cell(tbl, ri, ci, text, template_rpr=None):
    tr = tbl.findall(A+'tr')[ri]
    tc = tr.findall(A+'tc')[ci]
    ps = tc.findall('.//'+A+'p')
    set_para_text(ps[0], text, template_rpr)

def any_rpr(tbl):
    r = tbl.find('.//'+A+'r')
    return r.find(A+'rPr') if r is not None else None

# ---------------- 3. Chart parts (chart3 <- chart2) ----------------
shutil.copy(f"{U}/ppt/charts/chart2.xml", f"{U}/ppt/charts/chart3.xml")
shutil.copy(f"{U}/ppt/charts/colors2.xml", f"{U}/ppt/charts/colors3.xml")
shutil.copy(f"{U}/ppt/charts/style2.xml", f"{U}/ppt/charts/style3.xml")

ct = etree.parse(f"{U}/ppt/charts/chart3.xml", parser); ctr = ct.getroot()
ser = ctr.find('.//'+C+'ser')
# series name
ser.find('.//'+C+'tx//'+C+'v').text = "进取组合净值"
# categories (dates) and values (nav) -- full series incl 2019-12-31=1.0
cat_dates = [d.strftime('%Y-%m-%d') for d in dates]
val_nav   = [repr(v) for v in nav]
n = len(nav)
strcache = ser.find('.//'+C+'cat//'+C+'strCache')
ser.find('.//'+C+'cat//'+C+'f').text = f"结果页!$A$2:$A${n+1}"
for pt in strcache.findall(C+'pt'): strcache.remove(pt)
strcache.find(C+'ptCount').set('val', str(n))
for i,s in enumerate(cat_dates):
    pt = etree.SubElement(strcache, C+'pt'); pt.set('idx', str(i))
    etree.SubElement(pt, C+'v').text = s
numcache = ser.find('.//'+C+'val//'+C+'numCache')
ser.find('.//'+C+'val//'+C+'f').text = f"结果页!$D$2:$D${n+1}"
for pt in numcache.findall(C+'pt'): numcache.remove(pt)
numcache.find(C+'ptCount').set('val', str(n))
for i,s in enumerate(val_nav):
    pt = etree.SubElement(numcache, C+'pt'); pt.set('idx', str(i))
    etree.SubElement(pt, C+'v').text = s
ct.write(f"{U}/ppt/charts/chart3.xml", xml_declaration=True, encoding='UTF-8', standalone=True)

# chart3 rels (clone of chart2 rels: style3/colors3 + external link kept)
rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.microsoft.com/office/2011/relationships/chartStyle" Target="style3.xml"/><Relationship Id="rId2" Type="http://schemas.microsoft.com/office/2011/relationships/chartColorStyle" Target="colors3.xml"/><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/oleObject" Target="file:///C:\\Users\\liuxx02\\Desktop\\样板间version2.xlsx" TargetMode="External"/></Relationships>'''
open(f"{U}/ppt/charts/_rels/chart3.xml.rels","w",encoding="utf-8").write(rels)

# content types: add overrides for chart3/style3/colors3
ctp = f"{U}/[Content_Types].xml"
txt = open(ctp,encoding="utf-8").read()
add = ('<Override PartName="/ppt/charts/chart3.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
       '<Override PartName="/ppt/charts/style3.xml" ContentType="application/vnd.ms-office.chartstyle+xml"/>'
       '<Override PartName="/ppt/charts/colors3.xml" ContentType="application/vnd.ms-office.chartcolorstyle+xml"/>')
txt = txt.replace('</Types>', add+'</Types>')
open(ctp,"w",encoding="utf-8").write(txt)

# slide18 rels: add chart rel rId3 -> chart3.xml
rp = f"{U}/ppt/slides/_rels/slide18.xml.rels"
rt = etree.parse(rp, parser); rr = rt.getroot()
rel = etree.SubElement(rr, "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship")
rel.set("Id","rId3")
rel.set("Type","http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart")
rel.set("Target","../charts/chart3.xml")
rt.write(rp, xml_declaration=True, encoding='UTF-8', standalone=True)

# ---------------- 4. Slide 18 (<- slide 16) ----------------
s16 = etree.parse(f"{U}/ppt/slides/slide16.xml", parser).getroot()
s18t = etree.parse(f"{U}/ppt/slides/slide18.xml", parser); s18 = s18t.getroot()
tree16 = {first_cnvpr(sh).get('name'): sh for sh in s16.find('.//'+P+'spTree')
          if first_cnvpr(sh) is not None}
spTree18 = s18.find('.//'+P+'spTree')
nid = 101
for nm in ['矩形 4','表格 5','表格 6','图表 12']:
    el = copy.deepcopy(tree16[nm])
    first_cnvpr(el).set('id', str(nid)); nid += 1
    if nm == '图表 12':
        el.find('.//'+C+'chart').set(R+'id','rId3')
    spTree18.append(el)

# edit 表格5 (weights row): 50/50/0/0
t5 = get_table([x for x in spTree18 if first_cnvpr(x) is not None and first_cnvpr(x).get('name')=='表格 5'][0])
for ci,val in [(1,'50%'),(2,'50%'),(3,'0%'),(4,'0%')]:
    set_cell(t5, 1, ci, val)
# edit 表格6 (metrics)
t6 = get_table([x for x in spTree18 if first_cnvpr(x) is not None and first_cnvpr(x).get('name')=='表格 6'][0])
set_cell(t6, 0, 1, pct(cagr))
set_cell(t6, 1, 1, pct(arith_avg_annual))
set_cell(t6, 2, 1, pct(vol))
set_cell(t6, 3, 1, pct(-mdd))          # shown positive, like 30/70 "9.39%"
set_cell(t6, 4, 1, f"{sharpe:.2f}")
set_cell(t6, 5, 1, "6次")
# edit 矩形4 note
note_sp = [x for x in spTree18 if first_cnvpr(x) is not None and first_cnvpr(x).get('name')=='矩形 4'][0]
note_p = note_sp.find('.//'+A+'p')
note_txt = (f"注：数据来源为嘉实财富。策略历史表现不构成对未来业绩的承诺。"
            f"统计区间为2019年12月31日-2026年5月29日。 "
            f"2026/05/29的净值为{nav[-1]:.4f}。计算夏普比例时，无风险收益按0%计")
set_para_text(note_p, note_txt)
s18t.write(f"{U}/ppt/slides/slide18.xml", xml_declaration=True, encoding='UTF-8', standalone=True)

# ---------------- 5. Slide 19 (<- slide 17) ----------------
s17 = etree.parse(f"{U}/ppt/slides/slide17.xml", parser).getroot()
s19t = etree.parse(f"{U}/ppt/slides/slide19.xml", parser); s19 = s19t.getroot()
tree17 = {first_cnvpr(sh).get('name'): sh for sh in s17.find('.//'+P+'spTree')
          if first_cnvpr(sh) is not None}
spTree19 = s19.find('.//'+P+'spTree')
nid = 201
for nm in ['矩形 6','矩形 10','表格 3']:
    el = copy.deepcopy(tree17[nm])
    first_cnvpr(el).set('id', str(nid)); nid += 1
    spTree19.append(el)

# 矩形6: edit P0,P1; delete P2 (中证800 line)
box6 = [x for x in spTree19 if first_cnvpr(x) is not None and first_cnvpr(x).get('name')=='矩形 6'][0]
ps = box6.findall('.//'+A+'p')
p0 = (f"{len(full_years)}个完整年度有{pos_years}年为正收益。"
      f"最高年度收益{pct(annual[best_y]['ret'])}，最低年度收益{pct(annual[worst_y]['ret'])}，"
      f"年化收益率{pct(cagr)}。")
p1 = (f"{len(allm)}个完整月度录得{pos_cnt}个月({overall_wr*100:.1f}%)正收益。"
      f"最高月度收益{pct(max(allm))}，最低月度收益{pct(min(allm))}，"
      f"平均月度收益{pct(mean_m)}")
set_para_text(ps[0], p0)
set_para_text(ps[1], p1)
# remove the 3rd paragraph (its parent is the txBody)
ps[2].getparent().remove(ps[2])

# 表格3: fill monthly grid
t3 = get_table([x for x in spTree19 if first_cnvpr(x) is not None and first_cnvpr(x).get('name')=='表格 3'][0])
tpl = any_rpr(t3)
SP = "　"  # ideographic space for empty cells (matches 30/70)
for ri,y in enumerate(years, start=1):
    set_cell(t3, ri, 0, str(y), tpl)
    for m in range(1,13):
        txt = pct(mret[(y,m)]) if (y,m) in mret else SP
        set_cell(t3, ri, m, txt, tpl)
    set_cell(t3, ri, 13, pct(annual[y]['ret']), tpl)
    set_cell(t3, ri, 14, pct(annual[y]['wr']), tpl)
    set_cell(t3, ri, 15, pct(annual[y]['dd']), tpl)
# 成立以来 row (last data row)
last = len(years)+1
set_cell(t3, last, 0, "成立以来", tpl)
for m in range(1,13): set_cell(t3, last, m, SP, tpl)
set_cell(t3, last, 13, pct(cum), tpl)
set_cell(t3, last, 14, pct(overall_wr), tpl)
set_cell(t3, last, 15, pct(mdd), tpl)
s19t.write(f"{U}/ppt/slides/slide19.xml", xml_declaration=True, encoding='UTF-8', standalone=True)

print("BUILD OK. years rows:", years, "last row idx:", last, "table rows:", len(t3.findall(A+'tr')))
