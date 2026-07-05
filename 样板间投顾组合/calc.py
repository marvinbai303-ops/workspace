import openpyxl, numpy as np
wb = openpyxl.load_workbook('奇点+中债新综合+偏股净值数列.xlsx', data_only=True)
ws = wb['Sheet1']
dates=[]; qd=[]; zz=[]; pg=[]
for r in range(3,1553):
    dates.append(ws.cell(r,1).value)
    b=ws.cell(r,2).value
    qd.append(1.0 if isinstance(b,str) else b)
    zz.append(ws.cell(r,3).value)
    pg.append(ws.cell(r,4).value)
dates=np.array(dates); qd=np.array(qd,float); zz=np.array(zz,float); pg=np.array(pg,float)
qd/=qd[0]; zz/=zz[0]; pg/=pg[0]
def rets(s): return s[1:]/s[:-1]-1
rq=rets(qd); rz=rets(zz); rp=rets(pg)
years=(dates[-1]-dates[0]).days/365.25
n=len(rq)
print('years',round(years,3),'obs',n,'period',dates[0].date(),'to',dates[-1].date())

def metrics(series):
    r=rets(series)
    tot=series[-1]/series[0]-1
    ann=(series[-1]/series[0])**(1/years)-1
    vol=r.std(ddof=1)*np.sqrt(252)
    peak=np.maximum.accumulate(series)
    dd=series/peak-1
    mdd=dd.min()
    sharpe=ann/vol
    calmar=ann/abs(mdd)
    dr=r[r<0]
    dvol=dr.std(ddof=1)*np.sqrt(252)
    sortino=ann/dvol
    wr=(r>0).mean()
    return dict(tot=tot,ann=ann,vol=vol,mdd=mdd,sharpe=sharpe,calmar=calmar,sortino=sortino,wr=wr)

pgm=metrics(pg)
print('\n=== 偏股基金(930950) ===')
print('年化收益 {:.2f}%  年化波动 {:.2f}%  最大回撤 {:.2f}%  Sharpe {:.3f}  Calmar {:.3f}'.format(
    pgm['ann']*100,pgm['vol']*100,pgm['mdd']*100,pgm['sharpe'],pgm['calmar']))

results=[]
for w in range(0,101,5):
    sw=w/100
    pr=sw*rq+(1-sw)*rz
    nav=np.concatenate([[1.0],np.cumprod(1+pr)])
    m=metrics(nav)
    results.append((w,m))

print('\n股票% 债券% | 年化收益 | 年化波动 | 最大回撤 | Sharpe | Calmar | Sortino | 日胜率')
for w,m in results:
    print('{:3d}% {:3d}% | {:6.2f}% | {:6.2f}% | {:7.2f}% | {:.3f} | {:.3f} | {:.3f} | {:.1f}%'.format(
        w,100-w,m['ann']*100,m['vol']*100,m['mdd']*100,m['sharpe'],m['calmar'],m['sortino'],m['wr']*100))

print('\n=== 纯奇点组合 ===',{k:round(v,4) for k,v in metrics(qd).items()})
print('=== 纯中债新综合 ===',{k:round(v,4) for k,v in metrics(zz).items()})

# Find matching ratio by volatility and by maxdd (interpolate finer 1% steps)
fine=[]
for w in range(0,101):
    sw=w/100
    pr=sw*rq+(1-sw)*rz
    nav=np.concatenate([[1.0],np.cumprod(1+pr)])
    fine.append((w,metrics(nav)))
def closest(key,target):
    return min(fine,key=lambda x:abs(x[1][key]-target))
print('\n匹配偏股基金波动率的配比:', closest('vol',pgm['vol'])[0],'% 股票')
print('匹配偏股基金最大回撤的配比:', closest('mdd',pgm['mdd'])[0],'% 股票')
print('匹配偏股基金年化收益的配比:', closest('ann',pgm['ann'])[0],'% 股票')
EOF
