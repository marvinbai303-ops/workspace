# -*- coding: utf-8 -*-
import openpyxl, statistics, math, datetime, copy
from collections import OrderedDict
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

SRC="样板间数据整合0529.xlsx"
wb=openpyxl.load_workbook(SRC)

# ---------- 1. Load & rebase 进取 series to 2020-01-02=1.0 ----------
wj=openpyxl.load_workbook("进取组合回测净值序列.xlsx", data_only=True)["Sheet1"]
jrows=[]
for r in range(2,wj.max_row+1):
    d=wj.cell(r,1).value; v=wj.cell(r,2).value
    if d and v is not None:
        jrows.append((d.date() if hasattr(d,'date') else d, float(v)))
base_date=datetime.date(2020,1,2)
base_val=[v for d,v in jrows if d==base_date][0]
ser=[(d, v/base_val) for d,v in jrows if d>=base_date]   # rebased, 2020-01-02=1.0
dates=[d for d,_ in ser]; nav=[v for _,v in ser]
assert abs(nav[0]-1.0)<1e-12

# ---------- 2. Compute metrics (same methodology as 30-70) ----------
days=(dates[-1]-dates[0]).days
T=days/365
cum=nav[-1]/nav[0]-1
geo=(nav[-1]/nav[0])**(365/days)-1
arith=cum/T
rets=[nav[i]/nav[i-1]-1 for i in range(1,len(nav))]
vol=statistics.pstdev(rets)*math.sqrt(252)
peak=nav[0]; mdd=0.0
for v in nav:
    peak=max(peak,v); mdd=min(mdd, v/peak-1)
sharpe=geo/vol
# monthly
monthly=OrderedDict()
for d,v in ser: monthly[(d.year,d.month)]=v
mret=OrderedDict(); prev=nav[0]
for k in monthly: mret[k]=monthly[k]/prev-1; prev=monthly[k]
mall=list(mret.values())
mmean=statistics.mean(mall)
posm=sum(1 for x in mall if x>0); totm=len(mall); wr=posm/totm
# quarterly
q=OrderedDict()
for d,v in ser: q[(d.year,(d.month-1)//3)]=v
qret=[]; prev=nav[0]
for k in q: qret.append(q[k]/prev-1); prev=q[k]
qmean=statistics.mean(qret)
# annual
years=sorted(set(y for y,_ in monthly))
prev_ye=nav[0]; ann={}
for y in years:
    yv=[(d,v) for d,v in ser if d.year==y]; ye=yv[-1][1]
    aret=ye/prev_ye-1; prev_ye=ye
    ms=[mret[(y,m)] for m in range(1,13) if (y,m) in mret]
    pk=yv[0][1]; dd=0.0
    for _,v in yv:
        pk=max(pk,v); dd=min(dd, v/pk-1)
    ann[y]=dict(ret=aret, wr=sum(1 for x in ms if x>0)/len(ms), dd=dd)
print("进取(rebased): 结束净值=%.6f 累计=%.4f 几何=%.4f 算术=%.4f vol=%.4f mdd=%.4f sharpe=%.3f"%(nav[-1],cum,geo,arith,vol,mdd,sharpe))
print("月:mmean=%.6f 季:qmean=%.6f posm=%d totm=%d wr=%.4f"%(mmean,qmean,posm,totm,wr))
print("years:",years)

# ---------- 3. Add 进取 column to 原始组合净值 (rebased, aligned by date) ----------
ws0=wb["原始组合净值"]
def pdate(v):
    if isinstance(v,datetime.datetime): return v.date()
    if isinstance(v,datetime.date): return v
    return datetime.date.fromisoformat(str(v)[:10])
navmap={d:v for d,v in ser}
# header E1 styled like D1
ws0["E1"].value="进取组合"
ws0["E1"]._style=copy.copy(ws0["D1"]._style)
nmiss=0
for r in range(2, ws0.max_row+1):
    d=pdate(ws0.cell(r,1).value)
    cell=ws0.cell(r,5)
    cell._style=copy.copy(ws0.cell(r,4)._style)
    if d in navmap: cell.value=navmap[d]
    else: nmiss+=1
ws0.column_dimensions['E'].width=13.0
print("first-sheet 进取 col added; dates not found:",nmiss)

# ---------- 4. New sheet 进取组合 = copy of 30-70, overwrite data ----------
new=wb.copy_worksheet(wb["30-70"]); new.title="进取组合"
# reposition: after 30-70, before 指数点位
wb._sheets.remove(new)
idx=wb.sheetnames.index("30-70")+1
wb._sheets.insert(idx, new)

def setv(addr,val): new[addr].value=val   # keep style

# metrics block
setv("B2","2020-01-02"); setv("B3","2026-05-29"); setv("B4",1)
setv("B5",nav[-1]); setv("B6",cum); setv("B7",geo); setv("B8",arith)
setv("B9",vol); setv("B10",mdd); setv("B11",sharpe); setv("B12",mmean)
setv("B13",qmean); setv("B14",posm); setv("B15",totm); setv("B16",wr)

# monthly table: left block A20:M26, right block P20:AB26 + AC/AD/AE, 成立以来 row27, annual A30:D36
for i,y in enumerate(years):
    rr=20+i
    # left block
    new.cell(rr,1).value=y
    for m in range(1,13):
        new.cell(rr,2+(m-1)).value = mret.get((y,m), None)
    # right block
    new.cell(rr,16).value=y                      # P
    for m in range(1,13):
        new.cell(rr,17+(m-1)).value = mret.get((y,m), None)   # Q..AB
    new.cell(rr,29).value=ann[y]['ret']          # AC 年收益
    new.cell(rr,30).value=ann[y]['wr']           # AD 月胜率
    new.cell(rr,31).value=ann[y]['dd']           # AE 最大回撤
    # annual block A30:D36
    ar=30+i
    new.cell(ar,1).value=y
    new.cell(ar,2).value=ann[y]['ret']
    new.cell(ar,3).value=ann[y]['wr']
    new.cell(ar,4).value=ann[y]['dd']
# 成立以来 row 27 (P27 label already '成立以来' from copy)
new.cell(27,29).value=cum
new.cell(27,30).value=wr
new.cell(27,31).value=mdd

# ---------- 5. Heat map: clear static fills on Q20:AB26, add color-scale CF ----------
nofill=PatternFill(fill_type=None)
for rr in range(20,27):
    for c in range(17,29):  # Q..AB
        new.cell(rr,c).fill=nofill
rule=ColorScaleRule(start_type='min', start_color='FF63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFFFFFFF',
                    end_type='max', end_color='FFF8696B')
new.conditional_formatting.add('Q20:AB26', rule)

wb.save(SRC)
print("SAVED. sheets:", wb.sheetnames)
